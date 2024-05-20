import logging

from itemadapter import ItemAdapter
import openpyxl
from openpyxl.styles import Alignment
import pymysql
from scrapy.exceptions import DropItem


class BookSPipeline:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.append(('榜单', '小说', '作者', '链接'))

    def open_spider(self, spider):
        pass

    def close_spider(self, spider):
        # 保存Excel文件
        try:
            self.wb.save('book.xlsx')
        except PermissionError:
            print("Permission denied: 'book.xlsx'. Please close the file if it is open and try again.")

    def process_item(self, item, spider):
        adapter = ItemAdapter(item)
        bd = adapter.get('bd', [])
        itm_text = adapter.get('itm_text', [])
        itm = adapter.get('item', [])
        itm_href = adapter.get('itm_href', [])

        # 确保每个字段的值都是列表
        bd = self.ensure_list(bd)
        itm_text = self.ensure_list(itm_text)
        itm = self.ensure_list(itm)
        itm_href = self.ensure_list(itm_href)

        # 当前行
        current_row = self.ws.max_row + 1

        # 将字段的值逐行插入到Excel中
        for i in range(len(bd)):
            # 合并单元格并居中对齐
            self.ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 49, end_column=1)
            cell = self.ws.cell(row=current_row, column=1, value=bd[i])
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 插入对应的其他字段值
            for j in range(50):
                self.ws.cell(row=current_row + j, column=2,
                             value=itm_text[i * 50 + j] if (i * 50 + j) < len(itm_text) else '')
                self.ws.cell(row=current_row + j, column=3, value=itm[i * 50 + j] if (i * 50 + j) < len(itm) else '')
                self.ws.cell(row=current_row + j, column=4,
                             value='https://www.bqguu.cc' + itm_href[i * 50 + j] if (i * 50 + j) < len(
                                 itm_href) else '')

            # 更新当前行
            current_row += 50

        return item

    def ensure_list(self, value):
        if isinstance(value, list):
            return value
        return [value]


import logging
import pymysql
from scrapy.exceptions import DropItem


class DBSPipeline:
    def __init__(self):
        self.data = []
        self.list = []
        self.conn = pymysql.connect(host='localhost', port=3306, user='root', password='123456', database='book',
                                    charset='utf8')
        self.cursor = self.conn.cursor()
        self.logger = logging.getLogger(__name__)

    def open_spider(self, spider):
        pass

    def close_spider(self, spider):
        if len(self.data) > 0:
            self._write_to_db_()
        self.conn.close()

    def process_item(self, item, spider):
        try:
            bd_list = item.get('bd', [])
            itm_text_list = item.get('itm_text', [])
            itm_list = item.get('item', [])
            itm_href_list = item.get('itm_href', [])

            # Ensure all lists have the same length
            max_length = max(len(bd_list), len(itm_text_list), len(itm_list), len(itm_href_list))
            bd_list = bd_list[:max_length]
            itm_text_list = itm_text_list[:max_length]
            itm_list = itm_list[:max_length]
            itm_href_list = itm_href_list[:max_length]
            #榜单id
            bids = ['{:04d}'.format(1 + i) for i in range(len(bd_list))]
            #关联id  # 将每个元素重复50次
            extended_bids = [bid for bid in bids for _ in range(50)]
            # Zip lists into tuples
            href_list = ['https://www.bqguu.cc' + s for s in itm_href_list]
            # 压缩成列表
            data_tuples = list(zip(extended_bids, itm_text_list, itm_list, href_list))
            data_list = list(zip(bids, bd_list))

            self.data.extend(data_tuples)
            self.list.extend(data_list)

            if len(self.data) >= 100:
                self._write_to_db_()
                self.data.clear()
            return item
        except Exception as e:
            self.logger.error(f"Error while inserting item into database: {e}")
            raise DropItem(f"Error while inserting item into database: {e}")

    def _write_to_db_(self):
        try:
            self.cursor.executemany(
                'INSERT INTO book.book_hrf (bid,itm_text, itm_er, itm_href) VALUES (%s,%s, %s, %s)', self.data
            )
            self.cursor.executemany(
                'INSERT INTO book.book_bd (bid,book_bd) values (%s,%s)', self.list
            )
            self.conn.commit()
        except Exception as e:
            self.conn.rollback()
            self.logger.error(f"Error while writing to database: {e}")
            raise DropItem(f"Error while writing to database: {e}")
